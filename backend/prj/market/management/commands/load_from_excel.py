from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from prj.settings import DATA_DIR, BASE_DIR

CURRENT_DATA_DIR_PATH_FULL=fr'D:\GitHub\site_django_from_dima_herson\backend\init_data'

CURRENT_DATA_DIR = 'init_data'
CURRENT_DATA_DIR_PATH = fr'{BASE_DIR[:-3]}\{CURRENT_DATA_DIR}'                       
CURRENT_FILE_NAME = 'price.xlsx'

class Command(BaseCommand):


    def handle(self, *args, **options):
        # current_path = f'{CURRENT_DATA_DIR_PATH_FULL}\\{CURRENT_FILE_NAME}'
        current_path = f'{CURRENT_DATA_DIR_PATH}\\{CURRENT_FILE_NAME}'

        print('start importing in excel %s' % current_path)

        # wb = load_workbook(DATA_DIR+'\price.xlsx')
        wb = load_workbook(current_path)
        current_sheet = wb.get_sheet_names()[0]
        
        # print(wb.get_sheet_names())
        sheet = (wb.get_sheet_by_name(current_sheet))
        # print(sheet)
        # print(sheet.max_row)
        # print(sheet.max_column)

        for count in range(1, sheet.max_row+1):
            # print(sheet.cell(row=count, column=2).value)

            if not sheet.cell(row=count, column=1).value:
                print(sheet.cell(row=count, column=2).value)



